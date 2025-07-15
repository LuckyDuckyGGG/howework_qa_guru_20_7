import os
import shutil

from script_os import CURRENT_DIR
from script_os import TMP_DIR
import time
import requests
from selene import query
from selenium import webdriver
from selene.support.shared import browser
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pathlib
from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook
import csv

if not os.path.exists("homeworkDirectory"):
    os.mkdir("homeworkDirectory")
    print("Создал директорию для файлов ДЗ")
else:
    print("Директория для файлов ДЗ уже создана")

options = webdriver.ChromeOptions()
prefs = {
    "download.default_directory": TMP_DIR,
    "download.prompt_for_download": False
}
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
browser.config.driver = driver

browser.open("https://www.online-convert.com/ru/file-format/pdf")
download_pdf = browser.element("[title='Скачать example_complex.pdf']").get(query.attribute("href"))
content_pdf = requests.get(url=download_pdf).content
with open(os.path.join(TMP_DIR, "pdf_example.pdf"), "wb") as file:
    file.write(content_pdf)

browser.open("https://www.online-convert.com/ru/file-format/xlsx")
download_xlsx = browser.element('[title="Скачать example.xlsx"]').get(query.attribute("href"))
content_xlsx = requests.get(url=download_xlsx).content
with open(os.path.join(TMP_DIR, "xlsx_example.xlsx"), "wb") as file:
    file.write(content_xlsx)

browser.open("https://www.online-convert.com/ru/file-format/csv")
download_csv = browser.element('[title="Скачать example.csv"]').get(query.attribute("href"))
content_csv = requests.get(url=download_csv).content
with open(os.path.join(TMP_DIR, "csv_example.csv"), "wb") as file:
    file.write(content_csv)

directory_files = pathlib.Path(TMP_DIR)

with ZipFile("homework.zip", mode="w") as homework_archive:
    for file_path in directory_files.iterdir():
        homework_archive.write(file_path, arcname=file_path.name)

shutil.move("homework.zip", TMP_DIR)

def test_pdf_have_text():
    with ZipFile(os.path.join(TMP_DIR, "homework.zip")) as zip_file:
        with zip_file.open('pdf_example.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            pdf = reader.pages[0].extract_text()
            assert "PDF test file" in pdf

def test_xlsx_have_text():
    with ZipFile(os.path.join(TMP_DIR, "homework.zip")) as zip_file:
        with zip_file.open('xlsx_example.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            xlsx = (sheet.cell(row=3, column=2).value)
            assert "Provide example file for this type" == xlsx

def test_csv_have_text():
    with ZipFile(os.path.join(TMP_DIR, "homework.zip")) as zip_file:
        with zip_file.open('csv_example.csv', 'r') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            readercvs = list(csv.reader(content.splitlines()))
            row = readercvs[11]
            string = ",".join(row)
            assert "Month,Cecilia,Patty,Robert,Frank,total" in string











